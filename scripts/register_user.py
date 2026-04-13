#!/usr/bin/env python3
"""
register_user.py

Reads USER_DATA env var (JSON), appends to data/users.csv,
sorts rows by (team, anno_corso, cognome), and regenerates
data/users.xlsx with visual grouping and formatting.
"""

import csv
import json
import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Config ──────────────────────────────────────────────────────
CSV_PATH = Path("data/users.csv")
XLSX_PATH = Path("data/users.xlsx")

FIELDS = [
    "timestamp",
    "username",
    "full_name",
    "email",
    "team",
    "tesista",
    "anno_corso",
    "status",
]

SORT_KEYS = ["team", "anno_corso", "full_name"]


# ── Helpers ─────────────────────────────────────────────────────
def sort_key(row: dict) -> tuple:
    """
    Sort by: team (case-insensitive) → anno_corso (numeric, tesisti last) → full_name.
    Tesisti (anno_corso == '') sort after year 1-3 within the same team.
    """
    team = row.get("team", "").strip().lower()
    anno = row.get("anno_corso", "").strip()
    # Tesisti get a high sort value so they appear after year students
    anno_num = int(anno) if anno.isdigit() else 99
    name = row.get("full_name", "").strip().lower()
    return (team, anno_num, name)


def read_csv() -> list[dict]:
    """Read existing CSV or return empty list."""
    if not CSV_PATH.exists():
        return []
    with open(CSV_PATH, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return [row for row in reader]


def write_csv(rows: list[dict]) -> None:
    """Write sorted rows to CSV."""
    CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict]) -> None:
    """
    Generate a formatted XLSX with:
    - Header row (bold, dark background)
    - Rows grouped by team with alternating group colors
    - Team separator rows for visual clarity
    - Auto-sized columns
    """
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Utenti Perforce"

    # ── Styles ──
    header_font = Font(name="Aptos", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1C1C1E", end_color="1C1C1E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    header_border = Border(
        bottom=Side(style="thin", color="48484A"),
    )

    # Two alternating group fills (light backgrounds)
    group_fills = [
        PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid"),  # light gray
        PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),  # light blue
    ]

    separator_fill = PatternFill(start_color="2C2C2E", end_color="2C2C2E", fill_type="solid")
    separator_font = Font(name="Aptos", bold=True, color="F5F5F7", size=11)

    normal_font = Font(name="Aptos", size=11, color="1C1C1E")
    normal_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        bottom=Side(style="hair", color="D0D0D0"),
    )

    # Human-readable header labels
    header_labels = {
        "timestamp": "Data registrazione",
        "username": "Username",
        "full_name": "Nome completo",
        "email": "Email",
        "team": "Team / Progetto",
        "tesista": "Tesista",
        "anno_corso": "Anno di corso",
        "status": "Stato",
    }

    # ── Header row ──
    for col_idx, field in enumerate(FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_labels.get(field, field))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── Data rows with team grouping ──
    current_row = 2
    current_team = None
    group_color_idx = 0

    for row_data in rows:
        team = row_data.get("team", "").strip()

        # Insert team separator when team changes
        if team.lower() != (current_team or "").lower():
            if current_team is not None:
                # Empty separator row between groups
                for col_idx in range(1, len(FIELDS) + 1):
                    cell = ws.cell(row=current_row, column=col_idx, value="")
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                ws.row_dimensions[current_row].height = 6
                current_row += 1

            # Team header row
            cell = ws.cell(row=current_row, column=1, value=f"▸ {team}")
            cell.font = separator_font
            cell.fill = separator_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

            # Count members in this team
            team_count = sum(
                1 for r in rows if r.get("team", "").strip().lower() == team.lower()
            )
            cell_count = ws.cell(
                row=current_row,
                column=len(FIELDS),
                value=f"{team_count} membri",
            )
            cell_count.font = Font(name="Aptos", size=10, color="A1A1A6")
            cell_count.fill = separator_fill
            cell_count.alignment = Alignment(horizontal="right", vertical="center")

            # Fill remaining separator cells
            for col_idx in range(2, len(FIELDS)):
                ws.cell(row=current_row, column=col_idx).fill = separator_fill

            ws.row_dimensions[current_row].height = 24
            current_row += 1

            current_team = team
            group_color_idx = (group_color_idx + 1) % len(group_fills)

        # Data row
        fill = group_fills[group_color_idx]
        for col_idx, field in enumerate(FIELDS, start=1):
            value = row_data.get(field, "")

            # Format specific fields
            if field == "tesista":
                value = "Sì" if value.lower() == "yes" else "No"
            elif field == "anno_corso" and not value:
                value = "—"
            elif field == "timestamp" and value:
                # Show just date + time, not full ISO
                value = value.replace("T", " ").split(".")[0]
            elif field == "status":
                value = value.capitalize() if value else "Pending"

            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = normal_font
            cell.fill = fill
            cell.alignment = normal_align
            cell.border = thin_border

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # ── Auto-size columns ──
    for col_idx, field in enumerate(FIELDS, start=1):
        # Measure max content width
        max_len = len(header_labels.get(field, field))
        for row_idx in range(2, current_row):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))

        # Set column width with padding
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # ── Auto-filter on header ──
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}1"

    wb.save(XLSX_PATH)


# ── Main ────────────────────────────────────────────────────────
def main():
    # Parse input from environment
    raw = os.environ.get("USER_DATA")
    if not raw:
        print("ERROR: USER_DATA environment variable is empty")
        sys.exit(1)

    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON in USER_DATA: {e}")
        sys.exit(1)

    # Support both single user (dict) and batch (list)
    if isinstance(parsed, dict):
        new_users = [parsed]
    elif isinstance(parsed, list):
        new_users = parsed
    else:
        print("ERROR: USER_DATA must be a JSON object or array")
        sys.exit(1)

    if not new_users:
        print("ERROR: No users to register")
        sys.exit(1)

    print(f"Processing {len(new_users)} user(s)...")

    # Load existing data
    existing = read_csv()
    usernames = {r["username"].strip().lower() for r in existing}

    # Validate and append each user
    required = ["username", "full_name", "email", "team"]
    for i, new_user in enumerate(new_users, start=1):
        missing = [f for f in required if not new_user.get(f, "").strip()]
        if missing:
            print(f"WARNING: User {i} missing fields: {', '.join(missing)} — skipping")
            continue

        row = {field: new_user.get(field, "") for field in FIELDS}
        if not row["status"]:
            row["status"] = "pending"

        if row["username"].strip().lower() in usernames:
            print(f"WARNING: Username '{row['username']}' already exists. Adding with 'duplicate' status.")
            row["status"] = "duplicate"

        existing.append(row)
        usernames.add(row["username"].strip().lower())
        print(f"  + {row['username']} ({row['full_name']}) → {row['team']}")

    # Sort and write
    existing.sort(key=sort_key)

    write_csv(existing)
    print(f"\nCSV updated: {CSV_PATH} ({len(existing)} total users)")

    write_xlsx(existing)
    print(f"XLSX updated: {XLSX_PATH}")

    # Summary
    teams = {}
    for r in existing:
        t = r.get("team", "N/A").strip()
        teams[t] = teams.get(t, 0) + 1

    print("\n── Team summary ──")
    for t, count in sorted(teams.items()):
        print(f"  {t}: {count} members")


if __name__ == "__main__":
    main()
