#!/usr/bin/env python3
"""
xlsx_export.py

Genera l'XLSX formattato degli utenti, con raggruppamento per team.
La generazione girava dentro GitHub Actions e committava il file nella
repo: ora l'XLSX si produce solo in locale, su richiesta esplicita.

Modulo puro, senza CLI. Lo usano:
    python scripts/kv_status.py --xlsx utenti.xlsx
    python scripts/perforce_provision.py --export-xlsx utenti.xlsx

    from xlsx_export import write_xlsx
    write_xlsx(users, Path("utenti.xlsx"))
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FIELDS = [
    "timestamp", "username", "full_name", "email",
    "team", "tesista", "anno_corso", "status",
]

HEADER_LABELS = {
    "timestamp": "Data registrazione",
    "username": "Username",
    "full_name": "Nome completo",
    "email": "Email",
    "team": "Team / Progetto",
    "tesista": "Tesista",
    "anno_corso": "Anno di corso",
    "status": "Stato",
}


def sort_key(row: dict) -> tuple:
    """
    Ordina per: team (case-insensitive) → anno_corso (tesisti in coda) → full_name.
    I tesisti hanno anno_corso vuoto, quindi finiscono dopo gli anni 1-3.
    """
    team = (row.get("team") or "").strip().lower()
    anno = (row.get("anno_corso") or "").strip()
    anno_num = int(anno) if anno.isdigit() else 99
    name = (row.get("full_name") or "").strip().lower()
    return (team, anno_num, name)


def write_xlsx(rows: list[dict], path: Path) -> None:
    """
    XLSX con:
    - intestazione in grassetto su fondo scuro
    - righe raggruppate per team, colori alternati tra gruppi
    - riga separatrice per team con conteggio membri
    - colonne auto-dimensionate e auto-filtro
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    rows = sorted(rows, key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Utenti Perforce"

    # ── Stili ──
    header_font = Font(name="Aptos", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1C1C1E", end_color="1C1C1E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    header_border = Border(bottom=Side(style="thin", color="48484A"))

    group_fills = [
        PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid"),  # grigio chiaro
        PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),  # azzurro chiaro
    ]

    separator_fill = PatternFill(start_color="2C2C2E", end_color="2C2C2E", fill_type="solid")
    separator_font = Font(name="Aptos", bold=True, color="F5F5F7", size=11)

    normal_font = Font(name="Aptos", size=11, color="1C1C1E")
    normal_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(bottom=Side(style="hair", color="D0D0D0"))

    # ── Intestazione ──
    for col_idx, field in enumerate(FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=HEADER_LABELS.get(field, field))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── Righe dati, raggruppate per team ──
    current_row = 2
    current_team = None
    group_color_idx = 0

    for row_data in rows:
        team = (row_data.get("team") or "").strip()

        if team.lower() != (current_team or "").lower():
            if current_team is not None:
                # Riga vuota di stacco tra un gruppo e l'altro
                for col_idx in range(1, len(FIELDS) + 1):
                    cell = ws.cell(row=current_row, column=col_idx, value="")
                    cell.fill = PatternFill(
                        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                    )
                ws.row_dimensions[current_row].height = 6
                current_row += 1

            cell = ws.cell(row=current_row, column=1, value=f"▸ {team}")
            cell.font = separator_font
            cell.fill = separator_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

            team_count = sum(
                1 for r in rows if (r.get("team") or "").strip().lower() == team.lower()
            )
            cell_count = ws.cell(
                row=current_row, column=len(FIELDS), value=f"{team_count} membri"
            )
            cell_count.font = Font(name="Aptos", size=10, color="A1A1A6")
            cell_count.fill = separator_fill
            cell_count.alignment = Alignment(horizontal="right", vertical="center")

            for col_idx in range(2, len(FIELDS)):
                ws.cell(row=current_row, column=col_idx).fill = separator_fill

            ws.row_dimensions[current_row].height = 24
            current_row += 1

            current_team = team
            group_color_idx = (group_color_idx + 1) % len(group_fills)

        fill = group_fills[group_color_idx]
        for col_idx, field in enumerate(FIELDS, start=1):
            value = row_data.get(field, "") or ""

            if field == "tesista":
                value = "Sì" if value.lower() == "yes" else "No"
            elif field == "anno_corso" and not value:
                value = "—"
            elif field == "timestamp" and value:
                value = value.replace("T", " ").split(".")[0]
            elif field == "status":
                value = value.capitalize() if value else "Pending"

            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = normal_font
            cell.fill = fill
            cell.alignment = normal_align
            cell.border = thin_border

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # ── Larghezza colonne ──
    for col_idx, field in enumerate(FIELDS, start=1):
        max_len = len(HEADER_LABELS.get(field, field))
        for row_idx in range(2, current_row):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}1"

    wb.save(path)
